import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# ==== BÆ¯á»šC 1: Äá»c file nguá»“n ====
file_input = r"D:\1.python\1.Clonard-1\Cnard\Summary\Mirvish Village Billing Outline 01Sep-28Sep2025_v.xxSep2025.xlsx"
file_summary = r"D:\1.python\1.Clonard-1\Cnard\Summary\CGI Summary.xlsx"

# Row 4 lÃ  header â†’ bá» qua 3 dÃ²ng Ä‘áº§u
df = pd.read_excel(file_input, skiprows=3)

# Chuáº©n hÃ³a tÃªn cá»™t
df.columns = df.columns.str.strip()
df.columns = df.columns.str.replace(r"\s+", " ", regex=True)

print("âœ… Danh sÃ¡ch cá»™t:", df.columns.tolist())

# ==== BÆ¯á»šC 2: Gom nhÃ³m dá»¯ liá»‡u ====
rows = []

# Gom theo Date + Trade
grouped = df.groupby(["Date", "Trade"])

for (date, trade), g in grouped:
    # Regular
    reg_sum = g["Reg (Hrs)"].fillna(0).sum()
    if reg_sum > 0:
        rows.append({
            "Date": date,
            "Trade": f"{trade}: {len(g)}",
            "Hrs": reg_sum
        })

    # Overtime 1.5X
    ot15_sum = g["O / T 1.5X"].fillna(0).sum()
    if ot15_sum > 0:
        num_workers_ot = (g["O / T 1.5X"].fillna(0) > 0).sum()
        rows.append({
            "Date": date,
            "Trade": f"{trade}: {num_workers_ot}",
            "Hrs": ot15_sum
        })

    # Overtime 2X
    ot2_sum = g["O/T 2X"].fillna(0).sum()
    if ot2_sum > 0:
        num_workers_ot2 = (g["O/T 2X"].fillna(0) > 0).sum()
        rows.append({
            "Date": date,
            "Trade": f"{trade}: {num_workers_ot2} (OT2)",
            "Hrs": ot2_sum
        })

result = pd.DataFrame(rows)
print("âœ… Káº¿t quáº£ chuyá»ƒn Ä‘á»•i:\n", result.head())

# ==== BÆ¯á»šC 3: Ghi vÃ o file Summary cÃ³ sáºµn ====
wb = load_workbook(file_summary)
ws = wb.active  # hoáº·c ws = wb["Sheet1"]

# Thiáº¿t láº­p vá»‹ trÃ­ ghi
start_row = 11  # báº¯t Ä‘áº§u ghi tá»« row 11
col_date = 1    # cá»™t A
col_trade = 2   # cá»™t B
col_category = 3  # cá»™t C (má»›i thÃªm)
col_description = 4  # cá»™t D (má»›i thÃªm)
col_hrs = 5     # cá»™t E

# TÃ´ Ä‘á» cÃ¡c Ã´ dá»¯ liá»‡u má»›i
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

current_row = start_row
for date, group in result.groupby("Date"):
    first_row = current_row
    is_first_row_of_date = True
    
    for _, r in group.iterrows():
        # Ghi Date chá»‰ 1 láº§n (á»Ÿ dÃ²ng Ä‘áº§u tiÃªn cá»§a má»—i ngÃ y)
        if is_first_row_of_date:
            cell_date = ws.cell(row=current_row, column=col_date, value=date)
            cell_date.fill = red_fill  # TÃ´ Ä‘á» Ã´ ngÃ y

        # Ghi Trade + tÃ´ Ä‘á»
        cell_trade = ws.cell(row=current_row, column=col_trade, value=r["Trade"])
        cell_trade.fill = red_fill

        # Ghi cá»™t C (Category)
        if is_first_row_of_date:
            cell_category = ws.cell(row=current_row, column=col_category, value="General & Safety")
        else:
            cell_category = ws.cell(row=current_row, column=col_category, value="- ditto -")
        cell_category.fill = red_fill

        # Ghi cá»™t D (Description)
        if is_first_row_of_date:
            cell_description = ws.cell(row=current_row, column=col_description, value="Various")
        else:
            cell_description = ws.cell(row=current_row, column=col_description, value=' " ')
        cell_description.fill = red_fill

        # Ghi Hrs + tÃ´ Ä‘á»
        cell_hrs = ws.cell(row=current_row, column=col_hrs, value=r["Hrs"])
        cell_hrs.fill = red_fill

        current_row += 1
        is_first_row_of_date = False  # Sau dÃ²ng Ä‘áº§u tiÃªn, cÃ¡c dÃ²ng sau sáº½ dÃ¹ng "ditto" vÃ  " "

    # Náº¿u cÃ³ nhiá»u dÃ²ng cho cÃ¹ng 1 Date â†’ merge Date (tuá»³ chá»n, báº¡n cÃ³ thá»ƒ báº­t láº¡i náº¿u cáº§n)
    # if current_row - first_row > 1:
    #     ws.merge_cells(start_row=first_row, start_column=col_date,
    #                    end_row=current_row-1, end_column=col_date)

# LÆ°u láº¡i file
wb.save(file_summary)
print("ğŸ‰ ÄÃ£ ghi dá»¯ liá»‡u vÃ o Summary.xlsx thÃ nh cÃ´ng!")



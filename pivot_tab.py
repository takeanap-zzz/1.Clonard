# ktra ngay co match hay khong
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime
import re

def extract_dates_from_filename(filename):
    """
    Tách ngày bắt đầu và kết thúc từ tên file dạng:
    'Aug 11 - Aug 17 2025 Weekly Timesheet Input v6.xlsx'
    """
    pattern = r'([A-Za-z]+ \d{1,2})\s*-\s*([A-Za-z]+ \d{1,2})\s*(\d{4})'
    match = re.search(pattern, filename)
    if not match:
        return None, None
    start_str, end_str, year_str = match.groups()
    try:
        start_date = pd.to_datetime(f"{start_str} {year_str}").date()
        end_date = pd.to_datetime(f"{end_str} {year_str}").date()
        return start_date, end_date
    except:
        return None, None

def select_files():
    root = tk.Tk()
    root.withdraw()
    source_file = filedialog.askopenfilename(title="Chọn file nguồn", filetypes=[("Excel files", "*.xlsx")])
    target_files = filedialog.askopenfilenames(title="Chọn các file đích", filetypes=[("Excel files", "*.xlsx")])
    return source_file, target_files

def process_files(source_file, target_files):
    df = pd.read_excel(source_file, sheet_name="Sheet1")
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()
    
    df.rename(columns={
        'Sum of Reg (Hrs)': 'Reg',
        'Sum of O / T 1.5X': 'OT1',
        'Sum of O/T 2X': 'OT2',
        'Holiday': 'Holiday'
    }, inplace=True)
    
    if 'Date' not in df.columns:
        print("❌ Không tìm thấy cột 'Date' trong file nguồn.")
        return
    df['Date'] = pd.to_datetime(df['Date']).dt.date
    
    # ===== Bảng mapping Trade =====
    trade_mapping = {
        "foreman": "General Labour Foreman",
        "labour": "General Labour",
        "handyman": "Handyman",
        "handyman foreman": "Handyman Foreman",
        # thêm các trade khác ở đây nếu cần
    }
    
    for file_path in target_files:
        
        
        # Lấy ngày từ tên file
        start_date, end_date = extract_dates_from_filename(file_path)
        if not start_date or not end_date:
            print(f"❌ Bỏ qua {file_path}: Không thể lấy ngày từ tên file.")
            continue
        print(f"[i] Tuần: {start_date} đến {end_date}")
        
        # Lọc dữ liệu theo tuần
        weekly_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
        if weekly_df.empty:
            print(f"[!] Không có dữ liệu phù hợp với tuần này.")
            continue
        
        group_cols = ['Name', 'Trade']
        agg_dict = {
            'Reg': 'sum',
            'OT1': 'sum',
            'OT2': 'sum'
        }
        if 'Holiday' in df.columns:
            agg_dict['Holiday'] = 'first'
        summary = weekly_df.groupby(group_cols, as_index=False).agg(agg_dict)
        
        # Áp dụng mapping trade trước
        summary['MappedTrade'] = summary['Trade'].apply(lambda x: trade_mapping.get(str(x).strip().lower(), str(x).strip()))
        
        # Sắp xếp theo thứ tự ưu tiên
        def get_trade_priority(trade):
            # Kiểm tra có Handyman Foreman không
            has_handyman_foreman = any('handyman foreman' in str(t).lower() for t in summary['MappedTrade'])
            
            trade_lower = str(trade).lower()
            
            if has_handyman_foreman:
                # Thứ tự: General Labour Foreman, Handyman Foreman, Handyman, General Labour
                if 'general labour foreman' in trade_lower:
                    return 1
                elif 'handyman foreman' in trade_lower:
                    return 2
                elif 'handyman' in trade_lower:
                    return 3
                elif 'general labour' in trade_lower:
                    return 4
                else:
                    return 999  # Các trade khác xuống cuối
            else:
                # Thứ tự: General Labour Foreman, General Labour, Handyman
                if 'general labour foreman' in trade_lower:
                    return 1
                elif 'general labour' in trade_lower:
                    return 2
                elif 'handyman' in trade_lower:
                    return 3
                else:
                    return 999  # Các trade khác xuống cuối
        
        # Sắp xếp summary theo thứ tự ưu tiên
        summary['Priority'] = summary['MappedTrade'].apply(get_trade_priority)
        summary = summary.sort_values(['Priority', 'Name']).reset_index(drop=True)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        row = 33
        for _, record in summary.iterrows():
            # Debug info
            print(f"[DEBUG] Trade: '{record['MappedTrade']}' - Name: '{record['Name']}'")
            
            ws[f"C{row}"] = record['MappedTrade']
            ws[f"D{row}"] = record['Name']
            ws[f"E{row}"] = record['Reg']
            ws[f"F{row}"] = record['OT1']
            ws[f"G{row}"] = record['OT2']
            ws[f"H{row}"] = record.get('Holiday', "")
            row += 1
        
        wb.save(file_path)
        print(f"[✅] Đã cập nhật file: {file_path}")

if __name__ == "__main__":
    source_file, target_files = select_files()
    if source_file and target_files:
        process_files(source_file, target_files)
    else:
        print("❌ Bạn chưa chọn đủ file.")